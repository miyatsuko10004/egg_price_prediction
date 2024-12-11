import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from statsmodels.tsa.statespace.sarimax import SARIMAX
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import TimeSeriesSplit
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestRegressor
from scipy import stats
import openpyxl
from openpyxl.chart import LineChart, Reference
from datetime import datetime, timedelta
import warnings

# 警告メッセージの抑制
warnings.filterwarnings('ignore')

def load_and_preprocess_data(filepath):
    """データの読み込みと前処理を行う関数"""
    # データ読み込み
    data = pd.read_csv(filepath)
    
    # 日付列の変換
    data['Date'] = pd.to_datetime(data['Date'] + '/1', format='%Y/%m/%d')
    data.set_index('Date', inplace=True)
    
    # 数値列の変換と欠損値処理
    numeric_columns = [
        'egg_price', 'egg_production', 'egg_shipment', 'egg_recipt', 
        'chick_count', 'farmer_price', 'wholesale_price', 'retail_price', 
        'household_consumption_per_man', 'egg_import', 
        'chick_feed_shipment', 'chicken_feed_shipment', 'feed_price'
    ]
    
    for col in numeric_columns:
        if col in data.columns:
            data[col] = pd.to_numeric(data[col], errors='coerce')
    
    # 特徴量エンジニアリング
    data['feed_price_moving_avg'] = data['feed_price'].rolling(window=3).mean()
    data['egg_production_moving_avg'] = data['egg_production'].rolling(window=3).mean()
    
    # ラグ変数の作成
    for col in ['egg_price', 'feed_price', 'egg_production']:
        data[f'{col}_lag1'] = data[col].shift(1)
        data[f'{col}_lag2'] = data[col].shift(2)
    
    return data.dropna()

def create_feature_matrix(data):
    """特徴量マトリクスの作成"""
    features = [
        'feed_price', 'egg_production', 'chick_count', 
        'feed_price_moving_avg', 'egg_production_moving_avg',
        'feed_price_lag1', 'feed_price_lag2',
        'egg_production_lag1', 'egg_production_lag2',
        'egg_price_lag1', 'egg_price_lag2'
    ]
    X = data[features]
    y = data['egg_price']
    return X, y

def cross_validate_sarima(data):
    """SARIMA モデルのクロスバリデーション"""
    tscv = TimeSeriesSplit(n_splits=5)
    mse_scores = []
    
    for train_index, test_index in tscv.split(data):
        train = data.iloc[train_index]
        test = data.iloc[test_index]
        
        model = SARIMAX(train['egg_price'], order=(1,1,1), seasonal_order=(1,1,1,12))
        results = model.fit()
        
        predictions = results.forecast(steps=len(test))
        mse = mean_squared_error(test['egg_price'], predictions)
        mse_scores.append(mse)
    
    return np.mean(mse_scores)

def train_random_forest(X, y):
    """ランダムフォレストモデルのトレーニング"""
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    
    rf_model = RandomForestRegressor(n_estimators=100, random_state=42)
    rf_model.fit(X_scaled, y)
    
    return rf_model, scaler

def ensemble_forecast(sarima_model, rf_model, scaler, data, forecast_index):
    """アンサンブル予測"""
    # SARIMA 予測
    sarima_forecast = sarima_model.get_forecast(steps=len(forecast_index))
    sarima_mean = sarima_forecast.predicted_mean
    sarima_ci = sarima_forecast.conf_int()
    
    # ランダムフォレスト予測の準備
    last_data_point = data.iloc[-1]
    rf_features = [
        last_data_point['feed_price'], 
        last_data_point['egg_production'], 
        last_data_point['chick_count'], 
        last_data_point['feed_price_moving_avg'], 
        last_data_point['egg_production_moving_avg'],
        last_data_point['feed_price_lag1'], 
        last_data_point['feed_price_lag2'],
        last_data_point['egg_production_lag1'], 
        last_data_point['egg_production_lag2'],
        last_data_point['egg_price_lag1'], 
        last_data_point['egg_price_lag2']
    ]
    
    rf_features_scaled = scaler.transform([rf_features])
    rf_forecast = rf_model.predict(rf_features_scaled)
    
    # アンサンブル予測（単純平均）
    ensemble_forecast = (sarima_mean + rf_forecast[0]) / 2
    
    return ensemble_forecast, sarima_mean, sarima_ci

def create_excel_report(forecast_index, ensemble_forecast, sarima_mean, sarima_ci, metrics_df):
    """Excelレポートの作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "予測結果"
    
    # タイトル
    ws['A1'] = f"{forecast_index[0].strftime('%Y年%m月')}から{forecast_index[-1].strftime('%Y年%m月')}までの鶏卵価格予測結果"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    
    # 結果の作成
    results = pd.DataFrame({
        'Date': forecast_index.strftime('%Y-%m'),
        'Ensemble_Forecast': ensemble_forecast,
        'SARIMA_Forecast': sarima_mean.values,
        'Lower_CI': sarima_ci['lower egg_price'],
        'Upper_CI': sarima_ci['upper egg_price']
    })
    
    # 予測結果の書き込み
    ws.append(['Date', 'Ensemble_Forecast', 'SARIMA_Forecast', 'Lower_CI', 'Upper_CI'])
    for r in results.itertuples(index=False, name=None):
        ws.append(r)
    
    # メトリクスの追加
    ws.append([])
    ws.append(["モデル評価指標と説明"])
    for r in metrics_df.itertuples(index=False, name=None):
        ws.append(r)
    
    # チャートの作成
    chart = LineChart()
    chart.title = "鶏卵価格予測（アンサンブル）"
    chart.x_axis.title = "日付"
    chart.y_axis.title = "価格"
    
    data = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=len(results)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(results)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # X軸のラベル調整
    chart.x_axis.tickLblSkip = 3
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.textRotation = 45
    
    # チャートの追加
    ws.add_chart(chart, "G2")
    
    # Excelファイルの保存
    wb.save("enhanced_egg_price_forecast.xlsx")

def main():
    # データ読み込みと前処理
    data = load_and_preprocess_data("eggData.csv")
    
    # 特徴量マトリクスの作成
    X, y = create_feature_matrix(data)
    
    # SARIMA クロスバリデーション
    sarima_cv_score = cross_validate_sarima(data)
    print(f"SARIMA Cross-Validation MSE: {sarima_cv_score:.2f}")
    
    # ランダムフォレストモデルのトレーニング
    rf_model, scaler = train_random_forest(X, y)
    
    # SARIMAモデルの学習
    sarima_model = SARIMAX(data['egg_price'], order=(1,1,1), seasonal_order=(1,1,1,12))
    sarima_fit = sarima_model.fit()
    
    # 予測期間の設定
    today = datetime.now()
    start_date = today.replace(day=1)
    end_date = (today.replace(year=today.year + 2, month=4, day=1) - timedelta(days=1)).replace(day=31)
    forecast_index = pd.date_range(start=start_date, end=end_date, freq='MS')
    
    # アンサンブル予測
    ensemble_forecast, sarima_mean, sarima_ci = ensemble_forecast(
        sarima_fit, rf_model, scaler, data, forecast_index
    )
    
    # モデル評価指標
    mse = mean_squared_error(data['egg_price'][-12:], sarima_fit.fittedvalues[-12:])
    rmse = np.sqrt(mse)
    mae = mean_absolute_error(data['egg_price'][-12:], sarima_fit.fittedvalues[-12:])
    r2 = r2_score(data['egg_price'][-12:], sarima_fit.fittedvalues[-12:])
    
    # 評価指標の説明
    metrics_explanation = {
        'Metric': ['MSE', 'RMSE', 'MAE', 'R-squared', 'Confidence Interval'],
        'Value': [f'{mse:.2f}', f'{rmse:.2f}', f'{mae:.2f}', f'{r2:.2f}', '95%'],
        'Explanation': [
            '平均二乗誤差：予測誤差の二乗の平均。値が小さいほど良い。',
            '平方根平均二乗誤差：MSEの平方根。元のデータと同じ単位で誤差を表す。',
            '平均絶対誤差：予測誤差の絶対値の平均。',
            '決定係数：モデルの当てはまりの良さを0から1の間で表す。1に近いほど良い。',
            '予測値が95%の確率でこの区間内に収まると予想される範囲。'
        ]
    }
    metrics_df = pd.DataFrame(metrics_explanation)
    
    # 結果の出力
    print("\n予測結果と評価指標:")
    print(ensemble_forecast)
    print("\n評価指標:")
    print(metrics_df)
    
    # Excelレポートの作成
    create_excel_report(forecast_index, ensemble_forecast, sarima_mean, sarima_ci, metrics_df)
    
    print("\n結果が enhanced_egg_price_forecast.xlsx に出力されました。")

if __name__ == "__main__":
    main()
